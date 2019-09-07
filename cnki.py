import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import time

def getHTMLText(url):
   try:
r = requests.get(url, timeout = 30)
r.raise_for_status()
r.encoding = r.apparent_encoding
return r.text
except:
return "产生异常"

def infoWrite(sheet,soup,count):#共同代码提取出来
orgn = soup.find('div',{'class':'orgn'}).text
fund = soup.find('label',{'id':'catalog_FUND'})
if fund==None:#判断是否为空
fund = ""
else:
fund = fund.parent.text
fund = fund.replace(' ','')#去除字符串信息的无用空格
fund = fund.replace('\n','')
keywd = soup.find('label',{'id':'catalog_KEYWORD'})
if keywd==None:#判断是否为空
keywd=""
else:
keywd = keywd.parent.text
keywd = keywd.replace(' ','')#去除字符串信息的无用空格
keywd = keywd.replace('\n','')
print(keywd)
sheet.cell(row=count,column=4).value=orgn
sheet.cell(row=count,column=5).value=fund
sheet.cell(row=count,column=6).value=keywd

def getJournalInfos(start_url,end_url,sheet,count,book,path):  #爬取1994-2001年
for i in range(1994,2002):
if i in [1996,1997,1998]:   #判断是否为96-98年，因为网页格式有变化
for j in range(1,5):   #1994-2001年只有4个月
month = '0'+str(j) if len(str(j))==1 else str(j) #形成01、02这种格式数据
for k in range(19):
num = '.00'+str(k) if len(str(k))==1 else '.0'+str(k) #形成01、02这种格式数据
date = str(i)[-1] + month + num
dates = str(i) + month + num
                    time.sleep(4) #每次爬取休息4秒，反正被反扒识别封IP
                    url = start_url + date + end_url
                    html = getHTMLText(url)
                    soup = BeautifulSoup(html,'html.parser')
                    title = soup.find('title').get_text()
                    check = '知网节'
                    if check in title:
                            continue
                    else:
                        count += 1
                        name = soup.find('h2',{'class':'title'}).get_text()
                        author = soup.find('div',{'class':'author'})
                        if author==None:  #筛选掉不是论文的
                            continue
                        else:
                            author = author.text
                        infoWrite(sheet,soup,count)  #写入相应的数据
                        sheet.cell(row=count,column=1).value=dates
                        sheet.cell(row=count,column=2).value=name
                        sheet.cell(row=count,column=3).value=author
            book.save(path)                    
        else:
            for j in range(1,5):
                    month = '0'+str(j) if len(str(j))==1 else str(j) #形成01、02这种格式数据
                for k in range(19):
                        
                    num = '00'+str(k) if len(str(k))==1 else '0'+str(k) #形成01、02这种格式数据
                    date = str(i) + month + num
                    time.sleep(4)
                    url = start_url + date + end_url
                    html = getHTMLText(url)
                    soup = BeautifulSoup(html,'html.parser') 
                    title = soup.find('title')
                    if title==None:
                            continue
                    else:
                        title=title.get_text()
                    check = '知网节'
                    if check in title:
                            continue
                    else:
                        count += 1
                        name = soup.find('h2',{'class':'title'}).get_text()
                        author = soup.find('div',{'class':'author'})
                        if author==None:  #筛选掉不是论文的
                            continue
                        else:
                            author = author.text
                        infoWrite(sheet,soup,count)  #写入相应的数据
                        sheet.cell(row=count,column=1).value=date
                        sheet.cell(row=count,column=2).value=name
                        sheet.cell(row=count,column=3).value=author
            book.save(path)

def getJournalInfos_2(start_url,end_url,sheet,count,book,path):   #爬取2002-2018年
    for i in range(2002,2019):
            if i<2010:
                for j in range(1,7):  #2002-2010年只有6个月
                month = '0'+str(j) if len(str(j))==1 else str(j)
                for k in range(1,19):
                        if(len(str(k))==1):  #形成01、02这种格式数据
                        num = '00'+str(k)
                    else:
                        num = '0'+str(k)
                    date = str(i) + month + num
                    time.sleep(4)
                    url = start_url + date + end_url  #对目标论文的url进行拼接
                    html = getHTMLText(url)
                    soup = BeautifulSoup(html,'html.parser')
                    title = soup.find('title').get_text()
                    check = '知网节'
                    if check in title:
                            continue
                    else:
                        count += 1
                        name = soup.find('h2',{'class':'title'}).get_text()
                        author = soup.find('div',{'class':'author'})
                        if author==None:  #筛选掉不是论文的
                            continue
                        else:
                            author = author.text
                        infoWrite(sheet,soup,count)  #写入相应的数据
                        sheet.cell(row=count,column=1).value=date
                        sheet.cell(row=count,column=2).value=name
                        sheet.cell(row=count,column=3).value=author
            book.save(path)
        else:
            for j in range(1,13):  #2011-至今有6个月
                month = '0'+str(j) if len(str(j))==1 else str(j)
                for k in range(1,19):
                        if(len(str(k))==1):  #形成01、02这种格式的数据
                        num = '00'+str(k)
                    else:
                        num = '0'+str(k)
                    date = str(i) + month + num
                    time.sleep(4)
                    url = start_url + date + end_url
                    html = getHTMLText(url)
                    soup = BeautifulSoup(html,'html.parser')
                    title = soup.find('title').get_text()
                    check = '知网节'
                    if check in title:
                            continue
                    else:
                        count += 1
                        name = soup.find('h2',{'class':'title'}).get_text()
                        author = soup.find('div',{'class':'author'})
                        if author==None:  #筛选掉不是论文的情况
                            continue
                        else:
                            author = author.text
                        infoWrite(sheet,soup,count)  #写入相应的数据
                        sheet.cell(row=count,column=1).value=date
                        sheet.cell(row=count,column=2).value=name
                        sheet.cell(row=count,column=3).value=author
            book.save(path)
    
def main():
        start_url = 'http://kns.cnki.net/kcms/detail/detail.aspx?dbcode=CJFD&filename=JCYJ'
    end_url = '&dbname=CJFDLAST2018&uid=WEEvREcwSlJHSldRa1FhdXNXaEd1ZHpzdzVhUmY4cWRvWThpeGc5WGUyYz0=$9A4hF_YAuvQ5obgVAqNKPCYcEjKensW4IQMovwHtwkF4VYPoHbKxJw!!'
    count = 1
    book = Workbook()  #创建excel对象
    sheet = book.create_sheet("中国管理科学",0)
    title_name = ["时间","论文标题","作者","单位","基金","关键词"]
    for i in range(6):
            sheet.cell(row=1,column=i+1).value=title_name[i]   #标题栏信息
    path = '/Users/Administrator/Desktop/中国管理科学.xlsx'
    getJournalInfos(start_url,end_url,sheet,count,book,path)    #爬取1994-2001年
    getJournalInfos_2(start_url,end_url,sheet,count,book,path)  #爬取2002-2018年

main()